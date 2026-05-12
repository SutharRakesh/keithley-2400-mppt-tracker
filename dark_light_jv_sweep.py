from nicegui import ui, run
import pyvisa
import numpy as np
import pandas as pd
from pathlib import Path
from datetime import datetime
import time
import re
import os
import csv
import shutil
import plotly.graph_objects as go

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.cell.cell import MergedCell


# =========================
# User settings
# =========================

DEFAULT_GPIB_ADDRESS = 'GPIB0::15::INSTR'

BASE_DATA_FOLDER = Path(
    r'D:\4. Postdoc_RIT USA\OneDrive - rit.edu\Postdoc2_Darpa\Data\IV data'
)

WINDOW_WIDTH = 1360
WINDOW_HEIGHT = 850
LEFT_PANEL_WIDTH = 335

PLOT_WIDTH = 965
PLOT_HEIGHT = 710

DATA_SHEET_NAME = 'Data'
RAW_SHEET_NAME = 'Raw_Data'

DATETIME_ROW = 1
RUN_LABEL_ROW = 2
HEADER_ROW = 3
DATA_START_ROW = 4
SCAN_COL_WIDTH = 2

SAVE_RAW_CSV_EACH_POINT = True
RAW_CSV_FOLDER_NAME = 'raw_csv'
BACKUP_FOLDER_NAME = 'backups'

SAFE_SAVE_RETRIES = 5
SAFE_SAVE_WAIT_S = 1.0

# =========================
# Keithley backend settings
# =========================

K2400_TERMINALS = 'FRON'

# Set True only if Sense HI and Sense LO are really connected.
K2400_REMOTE_SENSE = True

K2400_AUTO_ZERO = True

K2400_DEFAULT_NPLC = 2.0
K2400_DEFAULT_CURRENT_RANGE_A = 0.00001   # 0 = auto current range
K2400_DEFAULT_AVERAGE_COUNT = 5
K2400_AVERAGE_TYPE = 'REP'

# Faster timeout. If Keithley hangs, app waits 5 s instead of 15 or 60 s.
K2400_TIMEOUT_MS = 10000

# Stop safely if current reaches 98% of compliance.
K2400_COMPLIANCE_STOP_FRACTION = 0.98

# Normal mode: no full reset and no debug queries.
# Turn these True only when troubleshooting.
K2400_USE_FULL_RESET_EACH_RUN = False
K2400_PRINT_DEBUG_QUERIES = False

PLOT_UPDATE_EVERY_N_POINTS = 1



# =========================
# File and data helpers
# =========================

def safe_name(text: str) -> str:
    text = str(text).strip()
    text = re.sub(r'[^A-Za-z0-9_.-]+', '_', text)
    text = text.strip('_')
    return text if text else 'sample'


def make_output_path(sample: str, iv_kind: str) -> Path:
    sample_clean = safe_name(sample)
    iv_clean = safe_name(iv_kind)
    BASE_DATA_FOLDER.mkdir(parents=True, exist_ok=True)
    return BASE_DATA_FOLDER / f'{sample_clean}_{iv_clean}_IV.xlsx'


def make_raw_csv_path(sample: str, iv_kind: str, run_label: str, scan_dt: datetime) -> Path:
    sample_clean = safe_name(sample)
    iv_clean = safe_name(iv_kind)
    timestamp = scan_dt.strftime('%Y%m%d_%H%M%S')
    folder = BASE_DATA_FOLDER / RAW_CSV_FOLDER_NAME
    folder.mkdir(parents=True, exist_ok=True)
    return folder / f'{sample_clean}_{iv_clean}_{safe_name(run_label)}_{timestamp}.csv'


def make_voltage_list(start_v: float, end_v: float, step_v: float) -> np.ndarray:
    if step_v <= 0:
        raise ValueError('Step size must be positive.')

    direction = 1 if end_v >= start_v else -1
    step = abs(step_v) * direction
    span = abs(end_v - start_v)
    n_points = int(np.floor(span / abs(step_v) + 1e-12)) + 1

    values = [start_v + i * step for i in range(n_points)]

    if direction > 0 and values[-1] < end_v - abs(step_v) * 1e-9:
        values.append(end_v)
    elif direction < 0 and values[-1] > end_v + abs(step_v) * 1e-9:
        values.append(end_v)

    values = np.array(values, dtype=float)
    values = np.round(values, 12)

    if len(values) < 2:
        raise ValueError('Voltage range and step size produce fewer than 2 points.')

    return values


def scan_datetime_text(scan_dt) -> str:
    if isinstance(scan_dt, datetime):
        return scan_dt.strftime('%Y-%m-%d %H:%M:%S')
    if scan_dt is None:
        return ''
    return str(scan_dt)


# =========================
# Raw CSV helpers
# =========================

RAW_FIELDNAMES = [
    'scan_datetime',
    'sample',
    'iv_kind',
    'run_id',
    'point_index',
    'V_set_V',
    'V_meas_V',
    'I_meas_A',
    'abs_I_A',
    'keithley_time_s',
    'smu_status',
    'raw_read',
]


def initialize_raw_csv(path: Path):
    if not SAVE_RAW_CSV_EACH_POINT:
        return

    path.parent.mkdir(parents=True, exist_ok=True)

    with open(path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=RAW_FIELDNAMES)
        writer.writeheader()


def append_point_to_raw_csv(path: Path, row: dict):
    if not SAVE_RAW_CSV_EACH_POINT:
        return

    with open(path, 'a', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=RAW_FIELDNAMES, extrasaction='ignore')
        writer.writerow(row)
        f.flush()
        os.fsync(f.fileno())


# =========================
# Workbook helpers
# =========================

def create_new_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = DATA_SHEET_NAME
    raw_ws = wb.create_sheet(RAW_SHEET_NAME)
    setup_raw_sheet(raw_ws)
    raw_ws.sheet_state = 'hidden'
    return wb


def setup_raw_sheet(raw_ws):
    if raw_ws.max_row == 1 and raw_ws.max_column == 1 and raw_ws.cell(row=1, column=1).value is None:
        for col, header in enumerate(RAW_FIELDNAMES, start=1):
            raw_ws.cell(row=1, column=col).value = header

    raw_ws.freeze_panes = 'A2'

    for col in range(1, len(RAW_FIELDNAMES) + 1):
        cell = raw_ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color='1D3557')
        cell.fill = PatternFill('solid', fgColor='E0F2FE')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        raw_ws.column_dimensions[get_column_letter(col)].width = 18

    raw_ws.column_dimensions['L'].width = 45


def ensure_sheets(wb):
    if DATA_SHEET_NAME not in wb.sheetnames:
        wb.create_sheet(DATA_SHEET_NAME, 0)

    if RAW_SHEET_NAME not in wb.sheetnames:
        raw_ws = wb.create_sheet(RAW_SHEET_NAME)
    else:
        raw_ws = wb[RAW_SHEET_NAME]

    setup_raw_sheet(raw_ws)
    raw_ws.sheet_state = 'hidden'
    return wb[DATA_SHEET_NAME], raw_ws


def create_or_load_workbook(path: Path):
    if not path.exists():
        return create_new_workbook()

    try:
        wb = load_workbook(path)
        ensure_sheets(wb)
        return wb
    except Exception:
        backup_folder = path.parent / BACKUP_FOLDER_NAME
        backup_folder.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        corrupt_backup = backup_folder / f'{path.stem}_CORRUPTED_{timestamp}{path.suffix}'

        try:
            shutil.copy2(path, corrupt_backup)
        except Exception:
            pass

        return create_new_workbook()


def safe_save_workbook(wb, path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    backup_folder = path.parent / BACKUP_FOLDER_NAME
    backup_folder.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
    temp_path = path.with_name(f'.{path.stem}_{timestamp}.tmp.xlsx')
    backup_path = backup_folder / f'{path.stem}_backup_{timestamp}.xlsx'

    last_error = None

    for attempt in range(1, SAFE_SAVE_RETRIES + 1):
        try:
            wb.save(temp_path)

            test_wb = load_workbook(temp_path, data_only=True, read_only=True)
            test_wb.close()

            if path.exists():
                shutil.copy2(path, backup_path)

            os.replace(temp_path, path)
            return

        except PermissionError as e:
            last_error = e
            time.sleep(SAFE_SAVE_WAIT_S)

        except Exception:
            try:
                if temp_path.exists():
                    temp_path.unlink()
            except Exception:
                pass
            raise

    try:
        if temp_path.exists():
            temp_path.unlink()
    except Exception:
        pass

    raise RuntimeError(
        f'Could not save Excel file after {SAFE_SAVE_RETRIES} attempts. '
        f'Close the workbook in Excel/OneDrive and try again.'
    ) from last_error


def get_next_run_label(path: Path, sample: str) -> str:
    sample_clean = safe_name(sample)

    if not path.exists():
        return f'{sample_clean}_1'

    try:
        wb = load_workbook(path, data_only=True, read_only=True)

        if DATA_SHEET_NAME not in wb.sheetnames:
            wb.close()
            return f'{sample_clean}_1'

        ws = wb[DATA_SHEET_NAME]
        run_numbers = []

        for col in range(1, ws.max_column + 1, SCAN_COL_WIDTH):
            value = ws.cell(row=RUN_LABEL_ROW, column=col).value

            if not value:
                continue

            match = re.fullmatch(rf'{re.escape(sample_clean)}_(\d+)', str(value))

            if match:
                run_numbers.append(int(match.group(1)))

        wb.close()

        if not run_numbers:
            return f'{sample_clean}_1'

        return f'{sample_clean}_{max(run_numbers) + 1}'

    except Exception:
        return f'{sample_clean}_1'

def increment_run_label(run_label: str, sample: str) -> str:
    sample_clean = safe_name(sample)

    match = re.fullmatch(rf'{re.escape(sample_clean)}_(\d+)', str(run_label))

    if match:
        next_number = int(match.group(1)) + 1
        return f'{sample_clean}_{next_number}'

    return f'{sample_clean}_1'


def load_existing_runs(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()

    try:
        wb = load_workbook(path, data_only=True, read_only=True)

        if DATA_SHEET_NAME not in wb.sheetnames:
            wb.close()
            return pd.DataFrame()

        ws = wb[DATA_SHEET_NAME]
        rows = []

        for col in range(1, ws.max_column + 1, SCAN_COL_WIDTH):
            scan_dt = ws.cell(row=DATETIME_ROW, column=col).value
            run_label = ws.cell(row=RUN_LABEL_ROW, column=col).value

            if not run_label:
                continue

            for row_idx in range(DATA_START_ROW, ws.max_row + 1):
                voltage = ws.cell(row=row_idx, column=col).value
                abs_current = ws.cell(row=row_idx, column=col + 1).value

                if voltage is None or abs_current is None:
                    continue

                try:
                    voltage = float(voltage)
                    abs_current = abs(float(abs_current))
                except Exception:
                    continue

                rows.append({
                    'run_id': str(run_label),
                    'scan_datetime': scan_datetime_text(scan_dt),
                    'V_meas_V': voltage,
                    'abs_I_A': abs_current,
                    'I_meas_A': np.nan,
                })

        wb.close()
        return pd.DataFrame(rows)

    except Exception:
        return pd.DataFrame()


# =========================
# Excel formatting and save
# =========================

def style_data_sheet(ws):
    thin = Side(style='thin', color='D7DEE8')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    date_fill = PatternFill('solid', fgColor='E0F2FE')
    run_fill = PatternFill('solid', fgColor='DCEBFF')
    header_fill = PatternFill('solid', fgColor='F4F8FF')

    ws.freeze_panes = 'A4'

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    for col in range(1, ws.max_column + 1):
        c_date = ws.cell(row=DATETIME_ROW, column=col)
        c_run = ws.cell(row=RUN_LABEL_ROW, column=col)
        c_header = ws.cell(row=HEADER_ROW, column=col)

        if not isinstance(c_date, MergedCell):
            c_date.font = Font(bold=True, size=10, color='075985')
            c_date.fill = date_fill
            c_date.alignment = Alignment(horizontal='center', vertical='center')
            c_date.number_format = 'yyyy-mm-dd hh:mm:ss'

        if not isinstance(c_run, MergedCell):
            c_run.font = Font(bold=True, size=12, color='1D3557')
            c_run.fill = run_fill
            c_run.alignment = Alignment(horizontal='center', vertical='center')

        if not isinstance(c_header, MergedCell):
            c_header.font = Font(bold=True, size=10, color='243B53')
            c_header.fill = header_fill
            c_header.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[DATETIME_ROW].height = 22
    ws.row_dimensions[RUN_LABEL_ROW].height = 24
    ws.row_dimensions[HEADER_ROW].height = 22

    for row in range(DATA_START_ROW, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)

            if isinstance(cell, MergedCell):
                continue

            if (col - 1) % SCAN_COL_WIDTH == 0:
                cell.number_format = '0.###############'
            else:
                cell.number_format = '0.###############E+00'


def find_data_max_row(ws):
    max_data_row = HEADER_ROW

    for col in range(1, ws.max_column + 1, SCAN_COL_WIDTH):
        run_label = ws.cell(row=RUN_LABEL_ROW, column=col).value

        if not run_label:
            continue

        for row in range(DATA_START_ROW, ws.max_row + 1):
            voltage = ws.cell(row=row, column=col).value
            abs_current = ws.cell(row=row, column=col + 1).value

            if voltage is None or abs_current is None:
                continue

            try:
                float(voltage)
                float(abs_current)
                max_data_row = max(max_data_row, row)
            except Exception:
                continue

    return max_data_row


def find_last_scan_col(ws):
    last_col = 0

    for col in range(1, ws.max_column + 1, SCAN_COL_WIDTH):
        run_label = ws.cell(row=RUN_LABEL_ROW, column=col).value
        scan_dt = ws.cell(row=DATETIME_ROW, column=col).value

        if run_label or scan_dt:
            last_col = col + SCAN_COL_WIDTH - 1

    return last_col


def rebuild_excel_plot_on_same_sheet(wb, sample: str, iv_kind: str):
    ws = wb[DATA_SHEET_NAME]

    ws._charts = []

    data_max_row = find_data_max_row(ws)

    if ws.max_row > data_max_row:
        ws.delete_rows(data_max_row + 1, ws.max_row - data_max_row)

    chart = ScatterChart()
    chart.title = f'{safe_name(sample)} {iv_kind} I-V'
    chart.x_axis.title = 'Voltage (V)'
    chart.y_axis.title = '|Current| (A)'
    chart.y_axis.scaling.logBase = 10
    chart.y_axis.numFmt = '0E+00'
    chart.x_axis.majorTickMark = 'out'
    chart.y_axis.majorTickMark = 'out'
    chart.y_axis.minorTickMark = 'none'
    chart.legend.position = 'b'
    chart.height = 16
    chart.width = 24
    chart.scatterStyle = 'lineMarker'

    last_data_col = find_last_scan_col(ws)

    for col in range(1, ws.max_column + 1, SCAN_COL_WIDTH):
        run_label = ws.cell(row=RUN_LABEL_ROW, column=col).value

        if not run_label:
            continue

        max_row = HEADER_ROW

        for row in range(DATA_START_ROW, data_max_row + 1):
            if ws.cell(row=row, column=col).value is not None:
                max_row = row

        if max_row < DATA_START_ROW:
            continue

        x_values = Reference(ws, min_col=col, min_row=DATA_START_ROW, max_row=max_row)
        y_values = Reference(ws, min_col=col + 1, min_row=DATA_START_ROW, max_row=max_row)

        series = Series(y_values, x_values, title=str(run_label))
        series.marker.symbol = 'circle'
        series.marker.size = 5
        chart.series.append(series)

    if last_data_col == 0:
        last_data_col = 2

    chart_start_col = last_data_col + 2
    chart_start_row = 10
    chart_cell = f'{get_column_letter(chart_start_col)}{chart_start_row}'
    ws.add_chart(chart, chart_cell)


def append_rows_to_raw_sheet(raw_ws, df_new: pd.DataFrame):
    setup_raw_sheet(raw_ws)

    for _, row in df_new.iterrows():
        raw_ws.append([
            row.get('scan_datetime', ''),
            row.get('sample', ''),
            row.get('iv_kind', ''),
            row.get('run_id', ''),
            int(row.get('point_index', 0)),
            float(row.get('V_set_V', np.nan)),
            float(row.get('V_meas_V', np.nan)),
            float(row.get('I_meas_A', np.nan)),
            float(row.get('abs_I_A', np.nan)),
            row.get('keithley_time_s', ''),
            row.get('smu_status', ''),
            row.get('raw_read', ''),
        ])

    raw_ws.sheet_state = 'hidden'


def save_run_to_xlsx(
    path: Path,
    sample: str,
    iv_kind: str,
    run_label: str,
    scan_dt: datetime,
    df_new: pd.DataFrame,
):
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = create_or_load_workbook(path)
    ws, raw_ws = ensure_sheets(wb)

    data_max_row = find_data_max_row(ws)

    if ws.max_row > data_max_row:
        ws.delete_rows(data_max_row + 1, ws.max_row - data_max_row)

    last_scan_col = find_last_scan_col(ws)
    start_col = 1 if last_scan_col == 0 else last_scan_col + 1

    ws.merge_cells(
        start_row=DATETIME_ROW,
        start_column=start_col,
        end_row=DATETIME_ROW,
        end_column=start_col + 1,
    )

    ws.merge_cells(
        start_row=RUN_LABEL_ROW,
        start_column=start_col,
        end_row=RUN_LABEL_ROW,
        end_column=start_col + 1,
    )

    ws.cell(row=DATETIME_ROW, column=start_col).value = scan_dt
    ws.cell(row=DATETIME_ROW, column=start_col).number_format = 'yyyy-mm-dd hh:mm:ss'
    ws.cell(row=RUN_LABEL_ROW, column=start_col).value = run_label

    ws.cell(row=HEADER_ROW, column=start_col).value = 'Voltage'
    ws.cell(row=HEADER_ROW, column=start_col + 1).value = 'Abs Current'

    for idx, row in df_new.iterrows():
        excel_row = int(idx) + DATA_START_ROW
        ws.cell(row=excel_row, column=start_col).value = float(row['V_meas_V'])
        ws.cell(row=excel_row, column=start_col + 1).value = float(row['abs_I_A'])

    append_rows_to_raw_sheet(raw_ws, df_new)

    style_data_sheet(ws)
    rebuild_excel_plot_on_same_sheet(wb, sample, iv_kind)

    safe_save_workbook(wb, path)


# =========================
# Keithley functions
# =========================

def clamp_float(value, low, high, name):
    value = float(value)
    if value < low or value > high:
        raise ValueError(f'{name} must be between {low} and {high}.')
    return value

def open_keithley(
    visa_address: str,
    start_v: float,
    end_v: float,
    compliance_mA: float,
    delay_s: float,
    nplc: float,
    current_range_A: float,
    average_count: int,
):
    rm = pyvisa.ResourceManager()
    inst = rm.open_resource(visa_address)

    inst.timeout = K2400_TIMEOUT_MS
    inst.read_termination = '\n'
    inst.write_termination = '\n'

    compliance_A = compliance_mA / 1000.0
    nplc = clamp_float(nplc, 0.01, 10.0, 'NPLC')
    delay_s = max(float(delay_s), 0.0)
    current_range_A = float(current_range_A)
    average_count = int(average_count)

    # Faster safe start.
    # Do not reset the whole instrument unless troubleshooting.
    try:
        inst.write(':OUTP OFF')
    except Exception:
        pass

    try:
        inst.write(':ABOR')
    except Exception:
        pass

    inst.write('*CLS')
    time.sleep(0.2)

    if K2400_USE_FULL_RESET_EACH_RUN:
        inst.write('*RST')
        time.sleep(1.0)
        inst.write('*CLS')

    inst.write(f':ROUT:TERM {K2400_TERMINALS}')

    inst.write(':SOUR:FUNC VOLT')
    inst.write(':SOUR:VOLT:MODE FIXED')
    inst.write(':SOUR:VOLT:LEV 0')

    v_range = max(abs(start_v), abs(end_v), 0.1) * 1.2
    inst.write(f':SOUR:VOLT:RANG {v_range:.6g}')

    inst.write(":SENS:FUNC 'CURR'")
    inst.write(f':SENS:CURR:PROT {compliance_A:.12g}')

    inst.write(':SOUR:DEL:AUTO OFF')
    inst.write(f':SOUR:DEL {delay_s:.6g}')

    inst.write(f':SENS:CURR:NPLC {nplc:.6g}')

    if K2400_AUTO_ZERO:
        inst.write(':SYST:AZER:STAT ON')
    else:
        inst.write(':SYST:AZER:STAT OFF')

    if current_range_A > 0:
        inst.write(':SENS:CURR:RANG:AUTO OFF')
        inst.write(f':SENS:CURR:RANG {current_range_A:.12g}')
    else:
        inst.write(':SENS:CURR:RANG:AUTO ON')

    if average_count >= 2:
        average_count = max(2, min(100, average_count))
        inst.write(f':SENS:AVER:TCON {K2400_AVERAGE_TYPE}')
        inst.write(f':SENS:AVER:COUN {average_count}')
        inst.write(':SENS:AVER ON')
    else:
        inst.write(':SENS:AVER OFF')

    if K2400_REMOTE_SENSE:
        inst.write(':SYST:RSEN ON')
    else:
        inst.write(':SYST:RSEN OFF')

    inst.write(':TRIG:COUN 1')
    inst.write(':ARM:COUN 1')

    inst.write(':FORM:DATA ASC')
    inst.write(':FORM:ELEM VOLT,CURR,TIME,STAT')
    inst.write(':OUTP OFF')

    # Debug queries are useful only when troubleshooting.
    # They slow startup because each query is a full GPIB round trip.
    # If one query has communication trouble, it can wait until timeout.
    if K2400_PRINT_DEBUG_QUERIES:
        try:
            print('Keithley compliance A:', inst.query(':SENS:CURR:PROT?').strip())
            print('Keithley current range auto:', inst.query(':SENS:CURR:RANG:AUTO?').strip())
            print('Keithley current range A:', inst.query(':SENS:CURR:RANG?').strip())
            print('Keithley remote sense:', inst.query(':SYST:RSEN?').strip())
            print('Keithley error:', inst.query(':SYST:ERR?').strip())
        except Exception as e:
            print(f'Could not query Keithley settings: {e}')

    return rm, inst

def measure_point(inst, voltage: float, max_retries: int = 0):
    """
    Measure one point safely.

    If the K2400 gets stuck waiting for a reading, the exception path turns output OFF,
    aborts the trigger/read, clears the bus if possible, and raises a clear error.
    """
    last_error = None

    for attempt in range(max_retries + 1):
        try:
            # Abort any previous incomplete trigger/read before starting the next point.
            try:
                inst.write(':ABOR')
            except Exception:
                pass

            inst.write(f':SOUR:VOLT:LEV {voltage:.12g}')

            raw = inst.query(':READ?').strip()
            values = [float(x) for x in raw.split(',')]

            if len(values) >= 4:
                v_meas = values[0]
                i_meas = values[1]
                t_rel = values[2]
                smu_status = values[3]
            elif len(values) == 2:
                v_meas = values[0]
                i_meas = values[1]
                t_rel = np.nan
                smu_status = np.nan
            elif len(values) == 1:
                v_meas = voltage
                i_meas = values[0]
                t_rel = np.nan
                smu_status = np.nan
            else:
                raise RuntimeError(f'Unexpected Keithley response: {raw}')

            return v_meas, i_meas, t_rel, smu_status, raw

        except pyvisa.errors.VisaIOError as e:
            last_error = e

            # Very important: do not leave the device/output in a stuck state.
            try:
                inst.write(':OUTP OFF')
            except Exception:
                pass

            try:
                inst.write(':ABOR')
                inst.write('*CLS')
            except Exception:
                pass

            try:
                inst.clear()
            except Exception:
                pass

            time.sleep(0.5)

    raise RuntimeError(
        f'Keithley did not respond at {voltage:.6g} V. '
        f'Output was turned OFF. Check compliance/current range/sample illumination.'
    ) from last_error


def close_keithley(inst, rm):
    try:
        if inst is not None:
            inst.write(':OUTP OFF')
    finally:
        try:
            if inst is not None:
                inst.close()
        finally:
            if rm is not None:
                rm.close()


# =========================
# Plot functions
# =========================

def choose_x_dtick(x_values):
    if not x_values:
        return 1

    x_min = min(x_values)
    x_max = max(x_values)
    span = abs(x_max - x_min)

    if span <= 0.5:
        return 0.1
    if span <= 1:
        return 0.2
    if span <= 3:
        return 0.5
    if span <= 10:
        return 1
    return 2


def build_plot(existing_df: pd.DataFrame, live_df: pd.DataFrame, sample_name: str = ''):
    fig = go.Figure()

    frames = []

    if existing_df is not None and not existing_df.empty:
        frames.append(existing_df)

    if live_df is not None and not live_df.empty:
        frames.append(live_df)

    all_abs_currents = []
    all_voltages = []

    if frames:
        all_df = pd.concat(frames, ignore_index=True)

        if 'run_id' in all_df.columns:
            for run_id, group in all_df.groupby('run_id'):
                group = group.copy()
                group['abs_I_A'] = group['abs_I_A'].abs()
                group.loc[group['abs_I_A'] <= 0, 'abs_I_A'] = np.nan

                valid_group = group.dropna(subset=['V_meas_V', 'abs_I_A'])

                if valid_group.empty:
                    continue

                all_abs_currents.extend(valid_group['abs_I_A'].tolist())
                all_voltages.extend(valid_group['V_meas_V'].tolist())

                fig.add_trace(
                    go.Scatter(
                        x=valid_group['V_meas_V'],
                        y=valid_group['abs_I_A'],
                        mode='lines+markers',
                        name=str(run_id),
                        line=dict(width=2),
                        marker=dict(size=6),
                        hovertemplate=(
                            '<b>%{fullData.name}</b><br>'
                            'V = %{x:.3f} V<br>'
                            '|I| = %{y:.3e} A<br>'
                            '<extra></extra>'
                        ),
                    )
                )

    y_range = None
    positive_currents = [i for i in all_abs_currents if i > 0]

    if positive_currents:
        y_min = min(positive_currents)
        y_max = max(positive_currents)
        y_min_exp = int(np.floor(np.log10(y_min))) - 1
        y_max_exp = int(np.ceil(np.log10(y_max))) + 1
        y_range = [y_min_exp, y_max_exp]

    x_dtick = choose_x_dtick(all_voltages)
    x_minor_dtick = x_dtick / 5
    title_sample = sample_name.strip() if sample_name.strip() else 'Sample'

    fig.update_layout(
        title=dict(
            text=f'{title_sample} - Semilog I-V',
            x=0.5,
            xanchor='center',
            font=dict(size=23, color='#0f172a'),
        ),
        template='plotly_white',
        width=PLOT_WIDTH,
        height=PLOT_HEIGHT,
        margin=dict(l=95, r=35, t=75, b=135),
        plot_bgcolor='rgba(240, 240, 240, 0.50)',
        paper_bgcolor='white',
        hovermode='closest',
        legend=dict(
            orientation='h',
            yanchor='top',
            y=-0.20,
            xanchor='center',
            x=0.5,
            font=dict(size=11),
            title=None,
        ),
    )

    fig.update_xaxes(
        title=dict(text='Voltage (V)', font=dict(size=18)),
        tickmode='linear',
        dtick=x_dtick,
        ticks='inside',
        ticklen=10,
        tickwidth=1.4,
        minor=dict(
            ticks='inside',
            dtick=x_minor_dtick,
            ticklen=5,
            tickwidth=1,
            showgrid=False,
        ),
        showline=True,
        linewidth=1.4,
        linecolor='black',
        mirror=True,
        showgrid=True,
        gridcolor='rgba(160, 160, 160, 0.50)',
        gridwidth=0.5,
        zeroline=True,
        zerolinewidth=1.4,
        zerolinecolor='black',
        tickfont=dict(size=13),
    )

    fig.update_yaxes(
        title=dict(text='|Current| (A)', font=dict(size=18)),
        type='log',
        range=y_range,
        tickmode='linear',
        dtick=1,
        exponentformat='power',
        showexponent='all',
        ticks='inside',
        ticklen=10,
        tickwidth=1.4,
        minor=dict(
            ticks='inside',
            ticklen=5,
            tickwidth=1,
            dtick='D1',
            showgrid=False,
        ),
        showline=True,
        linewidth=1.4,
        linecolor='black',
        mirror=True,
        showgrid=True,
        gridcolor='rgba(160, 160, 160, 0.50)',
        gridwidth=0.5,
        zeroline=False,
        tickfont=dict(size=13),
    )

    return fig


def refresh_plot(existing_df=None, live_df=None):
    if existing_df is None:
        existing_df = pd.DataFrame()

    if live_df is None:
        live_df = pd.DataFrame()

    try:
        sample = sample_name_input.value
    except Exception:
        sample = ''

    plot_element.figure = build_plot(existing_df, live_df, sample)
    plot_element.update()


# =========================
# App state and run logic
# =========================

stop_requested = False
current_run_df = pd.DataFrame()

loaded_existing_df = pd.DataFrame()
loaded_output_path = None
cached_next_run_label = None


def frontend_ready() -> bool:
    return all(
        name in globals()
        for name in [
            'plot_element',
            'file_label',
            'folder_label',
            'run_label_label',
            'raw_csv_label',
            'status_label',
            'sample_name_input',
            'iv_type_select',
        ]
    )

async def load_selected_sample(*_, clear_live: bool = True, quiet: bool = False):
    global loaded_existing_df
    global loaded_output_path
    global current_run_df
    global cached_next_run_label

    if not frontend_ready():
        return

    sample = str(sample_name_input.value or '').strip()
    iv_kind = str(iv_type_select.value or '').strip()

    if not sample:
        loaded_existing_df = pd.DataFrame()
        loaded_output_path = None
        cached_next_run_label = None

        if clear_live:
            current_run_df = pd.DataFrame()

        file_label.set_text('File: ---')
        raw_csv_label.set_text('Raw CSV: ---')
        run_label_label.set_text('Run label: ---')
        status_label.set_text('Enter sample name.')
        refresh_plot(pd.DataFrame(), pd.DataFrame())
        return

    output_path = make_output_path(sample, iv_kind)
    loaded_output_path = output_path

    file_label.set_text(f'File: {output_path.name}')
    folder_label.set_text(f'Folder: {output_path.parent}')

    if not quiet:
        status_label.set_text(f'Loading previous runs for {sample}...')

    # This is the slow Excel read. It now only happens when you click Load,
    # change sample name, change Dark/Light, or first run a not-loaded sample.
    loaded_existing_df = await run.io_bound(load_existing_runs, output_path)
    cached_next_run_label = await run.io_bound(get_next_run_label, output_path, sample)

    run_label_label.set_text(f'Next run label: {cached_next_run_label}')
    raw_csv_label.set_text('Raw CSV: ---')

    if clear_live:
        current_run_df = pd.DataFrame()

    refresh_plot(loaded_existing_df, current_run_df)

    if not quiet:
        if loaded_existing_df.empty:
            status_label.set_text(f'Ready. No previous data found for {output_path.name}.')
        else:
            n_runs = loaded_existing_df['run_id'].nunique()
            n_points = len(loaded_existing_df)
            status_label.set_text(
                f'Ready. Loaded {n_runs} previous run(s), {n_points} point(s).'
            )

def request_stop():
    global stop_requested
    stop_requested = True
    status_label.set_text('Stopping after current point...')


async def run_iv():
    global stop_requested
    global current_run_df
    global loaded_existing_df
    global loaded_output_path
    global cached_next_run_label

    stop_requested = False
    current_run_df = pd.DataFrame()

    run_button.disable()
    stop_button.enable()
    progress_bar.value = 0

    rm = None
    inst = None

    try:
        sample = sample_name_input.value.strip()

        if not sample:
            raise ValueError('Enter sample name.')

        iv_kind = iv_type_select.value

        start_v = float(start_voltage_input.value)
        end_v = float(end_voltage_input.value)
        step_v = float(step_voltage_input.value)
        compliance_mA = float(compliance_input.value)
        compliance_A = compliance_mA / 1000.0
        delay_s = float(delay_input.value)
        nplc = float(nplc_input.value)
        current_range_A = float(current_range_input.value)
        average_count = int(float(average_count_input.value))

        if compliance_mA <= 0:
            raise ValueError('Compliance must be positive.')

        if current_range_A > 0 and compliance_A > current_range_A:
            raise ValueError(
                'Compliance current is higher than the fixed current range. '
                'Increase current range or set Current range A = 0 for auto.'
            )

        voltages = make_voltage_list(start_v, end_v, step_v)


        output_path = make_output_path(sample, iv_kind)

        # Only load Excel if this sample/IV file is not already loaded.
        # After one load, Run I-V uses memory and simply appends the next run.
        if loaded_output_path != output_path or cached_next_run_label is None:
            await load_selected_sample(clear_live=True, quiet=True)

        existing_df = loaded_existing_df.copy()

        if cached_next_run_label is not None:
            run_label = cached_next_run_label
        else:
            run_label = await run.io_bound(get_next_run_label, output_path, sample)
            
        scan_dt = datetime.now()
        raw_csv_path = make_raw_csv_path(sample, iv_kind, run_label, scan_dt)

        initialize_raw_csv(raw_csv_path)
        refresh_plot(existing_df, pd.DataFrame())

        file_label.set_text(f'File: {output_path.name}')
        folder_label.set_text(f'Folder: {output_path.parent}')
        run_label_label.set_text(f'Run label: {run_label}')
        raw_csv_label.set_text(f'Raw CSV: {raw_csv_path.name if SAVE_RAW_CSV_EACH_POINT else "OFF"}')

        status_label.set_text('Opening Keithley 2400...')
        range_text = 'AUTO' if current_range_A <= 0 else f'{current_range_A:.1e} A'
        sense_text = '4-probe remote sense' if K2400_REMOTE_SENSE else '2-probe local sense'
        measure_mode_label.set_text(
            f'Mode: Front terminal | {sense_text} | NPLC {nplc:g} | Range {range_text} | '
            f'Compliance {compliance_A:.3e} A'
        )

        rm, inst = await run.io_bound(
            open_keithley,
            visa_address_input.value.strip(),
            start_v,
            end_v,
            compliance_mA,
            delay_s,
            nplc,
            current_range_A,
            average_count,
        )

        await run.io_bound(inst.write, f':SOUR:VOLT:LEV {voltages[0]:.12g}')
        await run.io_bound(inst.write, ':OUTP ON')

        rows = []
        compliance_hit = False

        for idx, v_set in enumerate(voltages):
            if stop_requested:
                break

            status_label.set_text(f'Measuring {idx + 1}/{len(voltages)} | Set {v_set:.3f} V')

            v_meas, i_meas, t_rel, smu_status, raw_read = await run.io_bound(
                measure_point,
                inst,
                float(v_set),
            )

            if abs(float(i_meas)) >= K2400_COMPLIANCE_STOP_FRACTION * compliance_A:
                compliance_hit = True
                status_label.set_text(
                    f'Compliance reached: I={i_meas:.3e} A, '
                    f'limit={compliance_A:.3e} A. Stopping safely.'
                )

            row = {
                'scan_datetime': scan_dt.strftime('%Y-%m-%d %H:%M:%S'),
                'sample': sample,
                'iv_kind': iv_kind,
                'run_id': run_label,
                'point_index': idx + 1,
                'V_set_V': float(v_set),
                'V_meas_V': float(v_meas),
                'I_meas_A': float(i_meas),
                'abs_I_A': abs(float(i_meas)),
                'keithley_time_s': float(t_rel) if not pd.isna(t_rel) else '',
                'smu_status': float(smu_status) if not pd.isna(smu_status) else '',
                'raw_read': raw_read,
            }

            append_point_to_raw_csv(raw_csv_path, row)

            rows.append(row)
            current_run_df = pd.DataFrame(rows)

            progress_bar.value = (idx + 1) / len(voltages)

            live_voltage_label.set_text(f'{v_meas:.3f} V')
            live_current_label.set_text(f'{i_meas:.3e} A')
            live_abs_current_label.set_text(f'{abs(i_meas):.3e} A')
            live_point_label.set_text(f'{idx + 1} / {len(voltages)}')

            if idx % PLOT_UPDATE_EVERY_N_POINTS == 0 or idx == len(voltages) - 1:
                refresh_plot(existing_df, current_run_df)

            if compliance_hit:
                await run.io_bound(inst.write, ':OUTP OFF')
                break

        await run.io_bound(inst.write, ':OUTP OFF')

        if rows:
            df_new = pd.DataFrame(rows)

            await run.io_bound(
                save_run_to_xlsx,
                output_path,
                sample,
                iv_kind,
                run_label,
                scan_dt,
                df_new,
            )

            if compliance_hit:
                status_label.set_text(
                    f'Compliance reached. Saved {run_label} to {output_path.name}. '
                    f'Increase Compliance mA if this current is expected.'
                )
            elif stop_requested:
                status_label.set_text(f'Stopped. Saved {run_label} to {output_path.name}')
            else:
                status_label.set_text(f'Completed. Saved {run_label} to {output_path.name}')

            loaded_existing_df = pd.concat(
                [existing_df, df_new],
                ignore_index=True,
            )

            current_run_df = pd.DataFrame()
            refresh_plot(loaded_existing_df, pd.DataFrame())

            try:
                reloaded_df = await run.io_bound(load_existing_runs, output_path)
                if not reloaded_df.empty:
                    loaded_existing_df = reloaded_df
                    refresh_plot(loaded_existing_df, pd.DataFrame())
            except Exception as reload_error:
                print(f'Could not reload saved Excel file: {reload_error}')

            cached_next_run_label = increment_run_label(run_label, sample)
            run_label_label.set_text(f'Next run label: {cached_next_run_label}')

            ui.notify('Data saved to Excel. Backup and raw CSV were also created.', type='positive')
        else:
            status_label.set_text('No data collected.')

    except Exception as e:
        status_label.set_text(f'Error: {e}')
        ui.notify(f'Error: {e}', type='negative')

    finally:
        try:
            if inst is not None or rm is not None:
                await run.io_bound(close_keithley, inst, rm)
        except Exception:
            pass

        run_button.enable()
        stop_button.disable()


# =========================
# Front end
# =========================

ui.colors(
    primary='#2563eb',
    secondary='#0891b2',
    positive='#16a34a',
    negative='#dc2626',
    warning='#f59e0b',
)

ui.add_head_html("""
<style>
body {
    overflow: hidden;
    background:
        radial-gradient(circle at top left, rgba(37,99,235,0.18), transparent 28%),
        radial-gradient(circle at bottom right, rgba(8,145,178,0.18), transparent 28%),
        linear-gradient(135deg, #eff6ff 0%, #f8fafc 48%, #ecfeff 100%);
    font-family: Inter, Arial, sans-serif;
}
.q-card {
    border-radius: 20px;
    box-shadow: 0 16px 40px rgba(15, 23, 42, 0.14);
}
.q-field {
    margin-bottom: 3px;
}
.left-panel {
    background: rgba(255,255,255,0.94);
    border: 1px solid rgba(37, 99, 235, 0.18);
    backdrop-filter: blur(10px);
    overflow-y: auto;
}
.plot-panel {
    background: rgba(255,255,255,0.96);
    border: 1px solid rgba(14, 116, 144, 0.18);
}
.title-chip {
    background: linear-gradient(90deg, #1d4ed8, #0891b2);
    color: white;
    padding: 12px 13px;
    border-radius: 16px;
}
.info-box {
    background: linear-gradient(135deg, #eff6ff, #ecfeff);
    border: 1px solid #bfdbfe;
    border-radius: 14px;
}
.status-box {
    background: linear-gradient(135deg, #f8fafc, #f0fdf4);
    border: 1px solid #bbf7d0;
    border-radius: 14px;
}
.metric-card {
    background: white;
    border: 1px solid #dbeafe;
    border-radius: 14px;
    padding: 8px;
    box-shadow: 0 6px 18px rgba(15, 23, 42, 0.06);
}
.metric-title {
    color: #64748b;
    font-size: 11px;
    font-weight: 700;
}
.metric-value {
    color: #0f172a;
    font-size: 15px;
    font-weight: 800;
}
.small-muted {
    color: #64748b;
    font-size: 11px;
}
</style>
""")

with ui.row().classes('w-screen h-screen no-wrap gap-3 p-3'):

    with ui.card().classes('h-full left-panel').style(
        f'width: {LEFT_PANEL_WIDTH}px; min-width: {LEFT_PANEL_WIDTH}px; padding: 12px;'
    ):
        with ui.column().classes('title-chip w-full gap-1'):
            ui.label('Dark / Light I-V').classes('text-xl font-bold')
            ui.label('Keithley 2400 SMU | dark / light JV').classes('text-xs')

        ui.separator()

        sample_name_input = ui.input(
            'Sample name',
            placeholder='',
            on_change=load_selected_sample,
        ).classes('w-full').props('dense outlined color=primary debounce=700')

        iv_type_select = ui.select(
            ['Dark', 'Light'],
            value='Dark',
            label='IV type',
            on_change=load_selected_sample,
        ).classes('w-full').props('dense outlined color=primary')

        ui.button(
            'Load sample data',
            on_click=load_selected_sample,
        ).props('color=secondary outline').classes('w-full')

        with ui.card().classes('w-full info-box').style('padding: 9px; box-shadow: none;'):
            measure_mode_label = ui.label('Mode: Front terminal | safe compliance mode').classes(
                'text-sm font-bold text-blue-900'
            )
            ui.label('Use Compliance mA = 1.0000 for 1 mA; 0.0010 is 1 µA.').classes('text-xs text-red-700')
            ui.label('Current range A = 0 means Keithley auto range.').classes('text-xs text-gray-600')
            ui.label('Excel Data sheet: Voltage | Abs Current').classes('text-xs text-gray-600')
            ui.label('Hidden Raw_Data sheet keeps signed current and metadata').classes('text-xs text-gray-600')
            ui.label('Raw CSV is saved point-by-point for crash recovery').classes('text-xs text-gray-600')

        with ui.grid(columns=2).classes('w-full gap-2'):
            start_voltage_input = ui.number(
                'Start V',
                value=-2.00,
                format='%.2f',
            ).classes('w-full').props('dense outlined color=primary')

            end_voltage_input = ui.number(
                'End V',
                value=5.00,
                format='%.2f',
            ).classes('w-full').props('dense outlined color=primary')

            step_voltage_input = ui.number(
                'Step V',
                value=0.05,
                min=0.001,
                format='%.3f',
            ).classes('w-full').props('dense outlined color=primary')

            compliance_input = ui.number(
                'Compliance mA',
                value=0.0010,   # 1.0000 mA = 1e-3 A. Do not use 0.0010 for 1 mA.
                min=0.0001,
                format='%.4f',
            ).classes('w-full').props('dense outlined color=primary')

        delay_input = ui.number(
            'Keithley delay per point (s)',
            value=0.5000,
            min=0.0001,
            format='%.4f',
        ).classes('w-full').props('dense outlined color=primary')

        with ui.grid(columns=2).classes('w-full gap-2'):
            nplc_input = ui.number(
                'NPLC',
                value=K2400_DEFAULT_NPLC,
                min=0.01,
                max=10.0,
                format='%.2f',
            ).classes('w-full').props('dense outlined color=primary')

            average_count_input = ui.number(
                'Avg count',
                value=K2400_DEFAULT_AVERAGE_COUNT,
                min=1,
                max=100,
                format='%.0f',
            ).classes('w-full').props('dense outlined color=primary')

        current_range_input = ui.number(
            'Current range A, 0 = auto',
            value=K2400_DEFAULT_CURRENT_RANGE_A,
            min=0,
            format='%.1e',
        ).classes('w-full').props('dense outlined color=primary')

        visa_address_input = ui.input(
            'VISA address',
            value=DEFAULT_GPIB_ADDRESS,
        ).classes('w-full').props('dense outlined color=primary')

        with ui.row().classes('w-full no-wrap gap-2'):
            run_button = ui.button(
                'Run I-V',
                on_click=run_iv,
            ).props('color=positive glossy').classes('w-1/2')

            stop_button = ui.button(
                'Stop',
                on_click=request_stop,
            ).props('color=negative glossy').classes('w-1/2')

        stop_button.disable()

        progress_bar = ui.linear_progress(value=0).props(
            'instant-feedback color=positive'
        ).classes('w-full mt-1')

        with ui.grid(columns=2).classes('w-full gap-2 mt-1'):
            with ui.column().classes('metric-card gap-0'):
                ui.label('Voltage').classes('metric-title')
                live_voltage_label = ui.label('--- V').classes('metric-value')

            with ui.column().classes('metric-card gap-0'):
                ui.label('Current').classes('metric-title')
                live_current_label = ui.label('--- A').classes('metric-value')

            with ui.column().classes('metric-card gap-0'):
                ui.label('|Current|').classes('metric-title')
                live_abs_current_label = ui.label('--- A').classes('metric-value')

            with ui.column().classes('metric-card gap-0'):
                ui.label('Point').classes('metric-title')
                live_point_label = ui.label('--- / ---').classes('metric-value')

        with ui.card().classes('w-full status-box mt-1').style('padding: 9px; box-shadow: none;'):
            status_label = ui.label('Ready.').classes('text-xs text-gray-800')
            run_label_label = ui.label('Run label: ---').classes('small-muted')
            file_label = ui.label('File: ---').classes('small-muted')
            raw_csv_label = ui.label('Raw CSV: ---').classes('small-muted')
            folder_label = ui.label(f'Folder: {BASE_DATA_FOLDER}').classes('small-muted')

    with ui.card().classes('h-full grow plot-panel').style('padding: 8px; overflow: hidden;'):
        with ui.row().classes('w-full justify-between items-center px-2'):
            ui.label('Live Semilog I-V Plot').classes('text-lg font-bold text-slate-800')

        with ui.row().classes('w-full justify-center'):
            plot_element = ui.plotly(
                build_plot(pd.DataFrame(), pd.DataFrame(), '')
            ).style(f'width: {PLOT_WIDTH}px; height: {PLOT_HEIGHT}px;')


ui.run(
    native=True,
    reload=False,
    title='Compact Dark / Light I-V',
    window_size=(WINDOW_WIDTH, WINDOW_HEIGHT),
    fullscreen=False,
)